"""Standalone one-off: French's Landfill daily diagnostic at phase threshold 0.10.

NOT part of the pipeline. Loads French's raw data, replays the six cleaning filters in order
at a fixed phase threshold of 0.10, attributes each producing interval to the first filter
that dropped it, and writes a formatted per-day Excel workbook to the output/ folder:

    output/<site>_daily_diagnostic.xlsx

One row per day present in the raw data, with columns:
    Date | Good/Bad | Clean Fraction | Producing Intervals | Surviving Intervals |
    Primary Kill Filter | Intervals Killed | Reason

The Reason column is a plain-English sentence built from that day's actual data — naming the
specific inverter string or meter station at fault and ending with why the day missed the 70%
clean threshold. Good days are green, bad days red; header frozen, columns auto-fit.

Run from the project root:   python french_diagnostic.py
"""

import contextlib
import io
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO))

from src import config, cleaners
from src.excel_loader import load_workbook

PHASE_THRESHOLD = 0.10
SITE_NEEDLE = "french"
OUTPUT_DIR = REPO / "output"

# Run the phase filter at 0.10 without depending on any pipeline changes: every filter reads
# this global, so setting it here pins the replay to 0.10.
config.MAX_PHASE_CURRENT_DEVIATION = PHASE_THRESHOLD

# Filters in pipeline order, paired with a short label used for kill attribution.
FILTER_STEPS = [
    ("Value Spikes",    cleaners.filter_value_spikes),
    ("Inverter Active", cleaners.filter_inverter_active),
    ("Inverter Comms",  cleaners.filter_inverter_comms),
    ("Meter Comms",     cleaners.filter_meter_comms),
    ("Phase Current",   cleaners.filter_phase_current),
    ("Gross Outliers",  cleaners.filter_gross_outliers),
]
_STEP_ORDER = {label: i for i, (label, _) in enumerate(FILTER_STEPS)}

# Short labels for the combined ("two filters roughly equal") sentence.
_SHORT_LABEL = {
    "Inverter Active": "inverter dropout",
    "Inverter Comms":  "inverter share fault",
    "Meter Comms":     "meter station fault",
    "Phase Current":   "phase imbalance",
    "Value Spikes":    "value spikes",
    "Gross Outliers":  "out-of-range efficiency",
}


def _short_inverter(col: str) -> str:
    """'Inverter (PS2) E' -> 'PS2 E'."""
    return col.replace("Inverter", "").replace("(", "").replace(")", "").strip()


def _meter_name(col: str) -> str:
    """'Generation Meter (PS2)' -> 'Generation Meter PS2'."""
    return col.replace("(", "").replace(")", "").strip()


def _ps_token(name: str) -> str:
    """Pull the 'PSn' station token from a column/stem, e.g. 'Generation Meter (PS1)' -> 'PS1'."""
    m = re.search(r"PS\s*\d+", name, re.IGNORECASE)
    return m.group().replace(" ", "") if m else name


def load_french() -> "object":
    """Load the workbook in data/raw/ and return French's SiteRecord."""
    xlsx = sorted((REPO / "data" / "raw").glob("*.xlsx"))
    if not xlsx:
        raise SystemExit("No .xlsx found in data/raw/")
    sink = io.StringIO()
    with contextlib.redirect_stdout(sink):  # mute the loader's progress chatter
        records = load_workbook(xlsx[0])
    for r in records:
        if SITE_NEEDLE in r.site_id.lower():
            return r
    raise SystemExit(f"No site matching '{SITE_NEEDLE}' in {xlsx[0].name}")


class _Signals:
    """Per-row diagnostic signals (indexed by raw row) used to name the day's culprit.

    These reconstruct, per interval, WHICH inverter string / meter station was at fault, so
    the Reason sentence can name it. They mirror the filters' own conditions (the meter/inverter
    share baselines are taken over all eligible producing rows, close enough for naming)."""

    def __init__(self, raw: pd.DataFrame, sid: str):
        self.inv_cols = cleaners._raw_inverter_cols(raw, sid)
        inv = raw[self.inv_cols].apply(pd.to_numeric, errors="coerce")
        self.inv_nan = inv.isna()
        self.inv_zero = (inv == 0)
        self.inv_bad = (inv <= 0) | self.inv_nan          # the filter_inverter_active condition

        # Meter-station share signals (multi-station sites only).
        self.m_cols = cleaners._meter_station_cols(raw, sid)
        if len(self.m_cols) >= 2:
            mst = raw[self.m_cols].apply(pd.to_numeric, errors="coerce")
            mtot = mst.sum(axis=1, skipna=True)
            mshare = mst.div(mtot.where(mtot > 0), axis=0)
            elig = pd.to_numeric(raw[config.COL_METER_PRODUCTION_KW], errors="coerce") \
                >= config.MIN_GEN_KW_FOR_SHARE_CHECK
            base = mshare[elig].median()
            rel = mshare.sub(base, axis=1).div(base.where(base > 0), axis=1)
            self.m_neg = mst < 0
            self.m_fail = self.m_neg | (rel < -config.MAX_METER_SHARE_DEVIATION)
        else:
            self.m_neg = self.m_fail = None

        # Per-phase imbalance signals, per station (per-row median reference, like the filter).
        self.phase_fail = {}   # stem -> bool Series (any leg > threshold off median, or negative)
        self.phase_neg = {}
        for stem, cols in cleaners._phase_current_groups(raw):
            ph = raw[cols].apply(pd.to_numeric, errors="coerce")
            med = ph.median(axis=1)
            rel = ph.sub(med, axis=0).abs().div(med.where(med > 0), axis=0)
            neg = (ph < 0).any(axis=1)
            self.phase_fail[stem] = (rel.max(axis=1) > PHASE_THRESHOLD) | neg
            self.phase_neg[stem] = neg


def _full_clause(filt: str, fidx: pd.Index, n_prod: int, sig: _Signals) -> str:
    """Descriptive clause naming the specific fault behind `filt` on the day's killed rows."""
    n = len(fidx)
    if filt == "Inverter Active":
        bsum = sig.inv_bad.loc[fidx].sum()
        dom = bsum.idxmax()
        cnt = int(bsum.max())
        nanc = int(sig.inv_nan.loc[fidx, dom].sum())
        zc = int(sig.inv_zero.loc[fidx, dom].sum())
        mode = "dropped out of telemetry" if nanc >= zc else "reported zero output"
        others = int((bsum > 0).sum()) - 1
        extra = f" (plus {others} other string{'s' if others != 1 else ''})" if others > 0 else ""
        return f"{_short_inverter(dom)} {mode} on {cnt} of {n_prod} intervals{extra}"
    if filt == "Inverter Comms":
        return (f"one or more inverter strings reported an abnormal share of generation on "
                f"{n} intervals (likely a comms or CT fault)")
    if filt == "Meter Comms":
        if sig.m_fail is not None:
            fsum = sig.m_fail.loc[fidx].sum()
            cnt = int(fsum.max())
            if cnt > 0:
                dom = fsum.idxmax()
                if int(sig.m_neg.loc[fidx, dom].sum()) >= cnt / 2:
                    return f"{_meter_name(dom)} read negative on {cnt} of {n_prod} intervals (meter station fault)"
                return (f"{_meter_name(dom)} share of meter output collapsed on {cnt} of "
                        f"{n_prod} intervals (meter station fault)")
        return f"a meter station's share deviated on {n} intervals (meter station fault)"
    if filt == "Phase Current":
        counts = {stem: int(s.loc[fidx].sum()) for stem, s in sig.phase_fail.items()}
        impl = sorted((s for s, c in counts.items() if c > 0), key=lambda s: -counts[s])
        names = "/".join(_ps_token(s) for s in impl) if impl else "the meter stations"
        return (f"phase-current imbalance exceeded the {PHASE_THRESHOLD * 100:.0f}% threshold on "
                f"{n} of {n_prod} intervals at {names}")
    if filt == "Value Spikes":
        return f"{n} of {n_prod} intervals carried physically impossible spike values"
    if filt == "Gross Outliers":
        return f"{n} of {n_prod} intervals fell outside the efficiency sanity band"
    return f"{n} of {n_prod} intervals were dropped by {filt}"


def _reason(idx_all: pd.Index, killed: pd.Series, n_prod: int, n_surv: int, sig: _Signals) -> str:
    """Plain-English explanation of the day's good/bad classification from its actual data."""
    if n_prod == 0:
        return "No production this day (all intervals below the nighttime threshold)"
    frac = n_surv / n_prod
    if frac >= config.GOOD_DAY_MIN_CLEAN_PCT:
        return f"Good day — {n_surv} of {n_prod} intervals clean across all filters"

    vc = killed.value_counts()
    items = sorted(vc.items(), key=lambda kv: (-kv[1], _STEP_ORDER.get(kv[0], 99)))
    top_f, top_n = items[0]
    sec_f, sec_n = items[1] if len(items) > 1 else (None, 0)
    pct = frac * 100

    # Two filters of comparable weight -> combined sentence.
    if sec_f is not None and sec_n >= 0.7 * top_n and sec_n >= 5:
        return (f"Bad day — {_SHORT_LABEL.get(top_f, top_f)} ({top_n} intervals) and "
                f"{_SHORT_LABEL.get(sec_f, sec_f)} ({sec_n} intervals) combined to drop below "
                f"the 70% clean threshold ({n_surv} of {n_prod} clean)")

    fidx = killed.index[killed.values == top_f]
    clause = _full_clause(top_f, fidx, n_prod, sig)
    return (f"Bad day — {clause}, leaving {n_surv} of {n_prod} intervals clean "
            f"({pct:.1f}%), below the 70% threshold")


def build_diagnostic(record) -> pd.DataFrame:
    """Replay the six filters at the fixed threshold and return the per-day diagnostic table."""
    raw = record.raw_df
    sid = record.site_id
    sig = _Signals(raw, sid)

    ts = pd.to_datetime(raw[config.COL_TIMESTAMP], errors="coerce")
    date = ts.dt.normalize()
    meter = pd.to_numeric(raw[config.COL_METER_PRODUCTION_KW], errors="coerce")
    prod = (meter > config.NIGHTTIME_KW_THRESHOLD).to_numpy()

    # Charge each removed row to the FIRST filter that dropped it (filters are pure).
    kill = pd.Series(pd.NA, index=raw.index, dtype="object")
    df = raw
    prev = raw.index
    sink = io.StringIO()
    with contextlib.redirect_stdout(sink):
        for label, fn in FILTER_STEPS:
            df = fn(df, sid)
            kill.loc[prev.difference(df.index)] = label
            prev = df.index

    work = pd.DataFrame({"date": date, "prod": prod, "kill": kill}, index=raw.index)
    work = work[work["date"].notna()]

    rows = []
    for day, sub in work.groupby("date"):
        prod_sub = sub[sub["prod"]]
        n_prod = len(prod_sub)
        killed = prod_sub["kill"].dropna()           # producing intervals removed, labelled
        n_surv = n_prod - len(killed)
        if n_prod == 0:
            rows.append({
                "Date": str(day.date()), "Good/Bad": "", "Clean Fraction": "",
                "Producing Intervals": 0, "Surviving Intervals": 0,
                "Primary Kill Filter": "", "Intervals Killed": 0,
                "Reason": _reason(prod_sub.index, killed, 0, 0, sig),
            })
            continue
        frac = n_surv / n_prod
        if len(killed):
            vc = killed.value_counts()
            primary, n_killed = str(vc.index[0]), int(vc.iloc[0])
        else:
            primary, n_killed = "None (all survived)", 0
        rows.append({
            "Date": str(day.date()),
            "Good/Bad": "Good" if frac >= config.GOOD_DAY_MIN_CLEAN_PCT else "Bad",
            "Clean Fraction": f"{frac * 100:.1f}%",
            "Producing Intervals": n_prod,
            "Surviving Intervals": n_surv,
            "Primary Kill Filter": primary,
            "Intervals Killed": n_killed,
            "Reason": _reason(prod_sub.index, killed, n_prod, n_surv, sig),
        })
    return pd.DataFrame(rows)


def write_workbook(frame: pd.DataFrame, site_id: str) -> Path:
    """Write the diagnostic frame to a formatted Excel workbook and return its path."""
    headers = ["Date", "Good/Bad", "Clean Fraction", "Producing Intervals",
               "Surviving Intervals", "Primary Kill Filter", "Intervals Killed", "Reason"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Diagnostic"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    for _, r in frame.iterrows():
        ws.append([r[h] for h in headers])
        fill = green if r["Good/Bad"] == "Good" else red if r["Good/Bad"] == "Bad" else None
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.freeze_panes = "A2"
    reason_col = get_column_letter(len(headers))     # last column holds the long sentence
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        if letter == reason_col:
            continue                                 # sized + wrapped separately below
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[letter].width = longest + 2
    ws.column_dimensions[reason_col].width = 95
    for cell in ws[reason_col]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / f"{site_id}_daily_diagnostic.xlsx"
    wb.save(path)
    return path


def main() -> None:
    record = load_french()
    frame = build_diagnostic(record)
    path = write_workbook(frame, record.site_id)

    good = int((frame["Good/Bad"] == "Good").sum())
    bad = int((frame["Good/Bad"] == "Bad").sum())
    print(f"{record.site_id} - phase threshold {PHASE_THRESHOLD:g}")
    print(f"  days written : {len(frame)}  (good {good}, bad {bad})")
    print(f"  workbook     : {path}")


if __name__ == "__main__":
    main()
