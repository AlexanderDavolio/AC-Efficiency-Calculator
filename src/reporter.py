"""Output layer: per-site cleaned CSVs and console tables."""

import calendar
import contextlib
import io
import re
import sys
import textwrap
from pathlib import Path
from typing import List

import pandas as pd
from tabulate import tabulate

from src.models import SiteRecord
from src import config
from src import calculator


OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

# When True, also print the Phase Imbalance Sensitivity diagnostic table (and compute it).
# Off by default so the report stays focused for a non-technical audience.
VERBOSE = False


# ── Good-day methodology + months kept out of reported efficiency ──────────────
# Reported efficiency is computed over GOOD DAYS only. A producing day (>=1 interval with
# raw meter > NIGHTTIME_KW_THRESHOLD) is "good" when at least config.GOOD_DAY_MIN_CLEAN_PCT
# of its production intervals survive every cleaning filter; otherwise it is a "bad" day and
# its intervals are excluded from the monthly and overall efficiency figures. The good/bad
# day counts are reported alongside the efficiency numbers (see summarise_by_month /
# summarise_site). The day's data still lives in the raw input and cleaned CSV — only the
# reported efficiency is restricted to good days.
#
# Two further reasons a whole month is excluded from every reported efficiency figure:
#   1. config: a SITE_CONFIGS excluded_months entry (e.g. a confirmed CT fault)
#   2. anomaly: the month's efficiency is statistically far below the site's own history
# Both are surfaced in the Data Gaps section. The data frames flow through two layers, each
# already restricted to good days:
#   _baseline_df   — config exclusions removed (the basis for anomaly detection)
#   _reportable_df — config AND anomaly exclusions removed (basis for all reported figures)
# Anomaly detection runs on _baseline_df (NOT _reportable_df) to avoid a circular dependency.

def _excluded_periods(site_id: str) -> set:
    """Monthly pd.Periods a site's config explicitly excludes (SITE_CONFIGS.excluded_months)."""
    cfg = config.SITE_CONFIGS.get(site_id)
    months = getattr(cfg, "excluded_months", None) if cfg else None
    if not months:
        return set()
    return {pd.Period(m, freq="M") for m in months}


def _drop_periods(df: pd.DataFrame, periods: set) -> pd.DataFrame:
    """Return df without rows whose timestamp falls in any of the given monthly Periods."""
    if not periods:
        return df
    per = pd.to_datetime(df[config.COL_TIMESTAMP], errors="coerce").dt.to_period("M")
    return df[~per.isin(periods)]


def _day_quality(record: SiteRecord) -> pd.DataFrame:
    """One row per producing calendar day, classified good/bad for the good-day methodology.

    Thin adapter over calculator.day_quality — the single source of truth shared with the
    adaptive search in cleaners.run_all_filters. A producing day has >=1 interval whose raw
    meter exceeds config.NIGHTTIME_KW_THRESHOLD (daytime-only sites: restricted to reporting
    intervals); it is "good" when at least config.GOOD_DAY_MIN_CLEAN_PCT of those intervals
    survived cleaning. Returns a DataFrame indexed by date with columns period, n_prod, n_clean,
    frac, good (empty frame with those columns when there are no producing days).
    """
    return calculator.day_quality(record.raw_df, record.enriched_df.index)


def _good_day_dates(record: SiteRecord) -> set:
    """Set of normalised Timestamps for the site's good days (see _day_quality)."""
    dq = _day_quality(record)
    return set(dq.index[dq["good"]])


def _keep_good_days(df: pd.DataFrame, good_dates: set) -> pd.DataFrame:
    """Return only the rows of df whose timestamp falls on one of the given good days."""
    if df.empty:
        return df
    day = pd.to_datetime(df[config.COL_TIMESTAMP], errors="coerce").dt.normalize()
    return df[day.isin(good_dates)]


def _monthly_weighted_efficiency(df: pd.DataFrame) -> pd.Series:
    """Energy-weighted efficiency (%) per month for months meeting the minimum-interval
    threshold, indexed by monthly Period. Matches the monthly table's avg_efficiency_pct,
    so anomaly stats are computed on exactly the figures that would be reported."""
    d = df[df[config.COL_EFFICIENCY_PCT].notna()]
    if d.empty:
        return pd.Series(dtype=float)
    per = pd.to_datetime(d[config.COL_TIMESTAMP], errors="coerce").dt.to_period("M")
    grp = d.groupby(per)
    counts = grp[config.COL_EFFICIENCY_PCT].count()
    meter = grp[config.COL_METER_PRODUCTION_KW].sum()
    inv = grp[config.COL_TOTAL_INVERTER_KW].sum()
    eff = 100.0 * meter / inv.where(inv > 0)
    return eff[counts >= config.MIN_CLEAN_INTERVALS_PER_MONTH]


def _anomalous_periods(record: SiteRecord) -> set:
    """Monthly Periods whose reported efficiency is statistically anomalous — more than
    config.ANOMALY_STD_THRESHOLD standard deviations BELOW the site's own median monthly
    efficiency. Computed on _baseline_df (config exclusions removed). Returns empty when
    there are fewer than config.ANOMALY_MIN_MONTHS months or no spread to judge."""
    eff = _monthly_weighted_efficiency(_baseline_df(record)).dropna()
    if len(eff) < config.ANOMALY_MIN_MONTHS:
        return set()
    median = eff.median()
    std = eff.std()  # sample std (ddof=1)
    if pd.isna(std) or std == 0:
        return set()
    cutoff = median - config.ANOMALY_STD_THRESHOLD * std
    return set(eff.index[eff < cutoff])


def _nonreportable_periods(record: SiteRecord) -> set:
    """All months kept out of reported efficiency: config-excluded ∪ statistically anomalous."""
    return _excluded_periods(record.site_id) | _anomalous_periods(record)


def _baseline_df(record: SiteRecord) -> pd.DataFrame:
    """enriched_df restricted to good days with config-excluded months removed — basis for
    anomaly detection."""
    df = _drop_periods(record.enriched_df, _excluded_periods(record.site_id))
    return _keep_good_days(df, _good_day_dates(record))


def _reportable_df(record: SiteRecord) -> pd.DataFrame:
    """enriched_df restricted to good days, with config-excluded AND statistically-anomalous
    months removed — the basis for ALL reported efficiency figures."""
    df = _drop_periods(record.enriched_df, _nonreportable_periods(record))
    return _keep_good_days(df, _good_day_dates(record))


# ── Energy-weighted roll-ups ──────────────────────────────────────────────────
# Reported efficiency and loss % are ALWAYS energy-weighted — the per-interval
# EFFICIENCY_PCT column exists only for diagnostics (min/max) and must never be
# averaged with .mean() for a headline figure. Weighting each interval equally
# would over-count low-energy dawn/dusk intervals; weighting by energy is the only
# physically correct roll-up. The interval-duration factor cancels, so summing the
# kW columns is equivalent to summing kWh.

def _weighted_sums(df: pd.DataFrame) -> tuple:
    """Return (sum_meter_kw, sum_inverter_kw) over intervals where efficiency is defined.

    EFFICIENCY_PCT is NaN exactly when the meter reading is missing or inverter total
    is non-positive, so masking on it pairs the two sums correctly. Summing each column
    independently would let a NaN-meter interval still contribute to the inverter sum,
    inflating the denominator and deflating the reported efficiency.
    """
    valid = df[df[config.COL_EFFICIENCY_PCT].notna()]
    return (
        valid[config.COL_METER_PRODUCTION_KW].sum(),
        valid[config.COL_TOTAL_INVERTER_KW].sum(),
    )


def _weighted_efficiency_pct(df: pd.DataFrame) -> float:
    """Energy-weighted efficiency: 100 * sum(meter_kw) / sum(inverter_kw)."""
    meter, inv_total = _weighted_sums(df)
    if inv_total <= 0:
        return float("nan")
    return 100.0 * meter / inv_total


def _weighted_loss_pct(df: pd.DataFrame) -> float:
    """Energy-weighted loss %: 100 * (sum(inverter_kw) - sum(meter_kw)) / sum(inverter_kw)."""
    meter, inv_total = _weighted_sums(df)
    if inv_total <= 0:
        return float("nan")
    return 100.0 * (inv_total - meter) / inv_total


# ── Per-site summaries ────────────────────────────────────────────────────────

def summarise_site(record: SiteRecord) -> dict:
    """Return a summary dict. Efficiency/loss figures use only reportable months (config
    exclusions removed); row counts and the date range reflect all clean data."""
    full = record.enriched_df               # all clean intervals — counts and date range
    rep = _reportable_df(record)            # good days, untrustworthy months removed — efficiency
    ts = full[config.COL_TIMESTAMP]

    # Good/bad day tally over reportable periods (matches the basis of the efficiency figure).
    dq = _day_quality(record)
    dq = dq[~dq["period"].isin(_nonreportable_periods(record))]
    good_days = int(dq["good"].sum())
    bad_days  = int(len(dq) - good_days)

    # Adaptive phase-threshold metadata recorded by cleaners.run_all_filters (on cleaned_df.attrs).
    # auto = it took more than one iteration, i.e. it was stepped up from the global default.
    attrs = getattr(record.cleaned_df, "attrs", None) or {}
    phase_threshold_used = attrs.get("phase_threshold_used")
    phase_threshold_auto = int(attrs.get("phase_threshold_iterations", 1)) > 1

    summary = {
        "site_name":          record.site_id,
        "phase_threshold_used": phase_threshold_used,
        "phase_threshold_auto": phase_threshold_auto,
        "total_raw_rows":     len(record.raw_df),
        "clean_rows":         len(full),
        "clean_row_pct":      round(len(full) / len(record.raw_df) * 100, 2),
        "good_days":          good_days,
        "bad_days":           bad_days,
        "avg_efficiency_pct": round(_weighted_efficiency_pct(rep), 3),
        "min_efficiency_pct": round(rep[config.COL_EFFICIENCY_PCT].min(), 3),
        "max_efficiency_pct": round(rep[config.COL_EFFICIENCY_PCT].max(), 3),
        "avg_loss_delta_kw":     round(rep[config.COL_LOSS_DELTA_KW].mean(), 3),
        "avg_loss_pct":          round(_weighted_loss_pct(rep), 3),
        "total_energy_lost_kwh": round(rep[config.COL_ENERGY_LOST_KWH].sum(), 1),
        "date_range_start":      ts.min().date() if not ts.empty else None,
        "date_range_end":        ts.max().date() if not ts.empty else None,
    }

    record.summary = summary
    return summary


def summarise_by_month(record: SiteRecord) -> pd.DataFrame:
    """Return a DataFrame with one row per (year, month) using reportable clean intervals
    (config-excluded months removed; months below the minimum-interval threshold dropped)."""
    df = _reportable_df(record).copy()
    # Restrict to intervals where efficiency is defined (meter present, inverter > 0)
    # so the energy-weighted meter/inverter sums below stay paired. See _weighted_sums.
    df = df[df[config.COL_EFFICIENCY_PCT].notna()]
    ts = pd.to_datetime(df[config.COL_TIMESTAMP])
    df["_year"]  = ts.dt.year
    df["_month"] = ts.dt.month

    agg = (
        df.groupby(["_year", "_month"])
        .agg(
            row_count=(config.COL_EFFICIENCY_PCT, "count"),
            min_efficiency_pct=(config.COL_EFFICIENCY_PCT, "min"),
            max_efficiency_pct=(config.COL_EFFICIENCY_PCT, "max"),
            avg_loss_delta_kw=(config.COL_LOSS_DELTA_KW, "mean"),
            total_loss_delta_kw=(config.COL_LOSS_DELTA_KW, "sum"),
            total_energy_lost_kwh=(config.COL_ENERGY_LOST_KWH, "sum"),
            _sum_meter_kw=(config.COL_METER_PRODUCTION_KW, "sum"),
            _sum_inverter_kw=(config.COL_TOTAL_INVERTER_KW, "sum"),
        )
        .reset_index()
        .sort_values(["_year", "_month"])
        .reset_index(drop=True)
    )

    # Drop months with too few clean intervals to yield a meaningful figure — they are
    # surfaced in the Data Gaps section instead (see summarise_gap_months). This keeps a
    # handful of timing-jittered or CT-dropout-driven rows from being reported as a number.
    agg = agg[agg["row_count"] >= config.MIN_CLEAN_INTERVALS_PER_MONTH].reset_index(drop=True)

    # Energy-weighted efficiency and loss % per month (not a mean of per-interval ratios).
    inv_pos = agg["_sum_inverter_kw"].where(agg["_sum_inverter_kw"] > 0)
    agg["avg_efficiency_pct"] = 100.0 * agg["_sum_meter_kw"] / inv_pos
    agg["avg_loss_pct"] = 100.0 * (agg["_sum_inverter_kw"] - agg["_sum_meter_kw"]) / inv_pos

    if agg.empty:
        return pd.DataFrame(columns=["Month", "Good Days", "Bad Days", "Valid Readings",
                                     "Average Efficiency (%)", "Min Efficiency (%)",
                                     "Max Efficiency (%)", "Average Power Loss (kW)",
                                     "Loss % of Output", "Total Power Loss (kW)",
                                     "Total Energy Lost (kWh)"])

    float_cols = ["avg_efficiency_pct", "min_efficiency_pct", "max_efficiency_pct",
                  "avg_loss_delta_kw", "total_loss_delta_kw"]
    agg[float_cols] = agg[float_cols].round(3)
    agg["avg_loss_pct"] = agg["avg_loss_pct"].round(3)
    agg["total_energy_lost_kwh"] = agg["total_energy_lost_kwh"].round(1)

    month_labels = agg.apply(
        lambda r: f"{calendar.month_abbr[int(r['_month'])]} {int(r['_year'])}", axis=1
    )

    # Good/bad day counts per reported month (same good-day basis as the efficiency above).
    dq = _day_quality(record)
    dq = dq[~dq["period"].isin(_nonreportable_periods(record))]
    good_by_period = dq.groupby("period")["good"].sum()
    prod_by_period = dq.groupby("period").size()
    periods = agg.apply(
        lambda r: pd.Period(year=int(r["_year"]), month=int(r["_month"]), freq="M"), axis=1
    )
    good_days = periods.map(lambda p: int(good_by_period.get(p, 0)))
    bad_days  = periods.map(lambda p: int(prod_by_period.get(p, 0)) - int(good_by_period.get(p, 0)))

    return pd.DataFrame({
        "Month":                    month_labels,
        "Good Days":                good_days,
        "Bad Days":                 bad_days,
        "Valid Readings":           agg["row_count"],
        "Average Efficiency (%)":   agg["avg_efficiency_pct"],
        "Min Efficiency (%)":       agg["min_efficiency_pct"],
        "Max Efficiency (%)":       agg["max_efficiency_pct"],
        "Average Power Loss (kW)":  agg["avg_loss_delta_kw"],
        "Loss % of Output":         agg["avg_loss_pct"],
        "Total Power Loss (kW)":    agg["total_loss_delta_kw"],
        "Total Energy Lost (kWh)":  agg["total_energy_lost_kwh"],
    })


def summarise_by_time_bucket(record: SiteRecord) -> pd.DataFrame:
    """Return a DataFrame with one row per time bucket (excluding 'Other')."""
    df = _reportable_df(record)
    # Keep only efficiency-defined intervals so the energy-weighted sums stay paired.
    df = df[(df[config.COL_TIME_BUCKET] != "Other") & df[config.COL_EFFICIENCY_PCT].notna()]

    agg = (
        df.groupby(config.COL_TIME_BUCKET)
        .agg(
            row_count=(config.COL_EFFICIENCY_PCT, "count"),
            avg_loss_delta_kw=(config.COL_LOSS_DELTA_KW, "mean"),
            _sum_meter_kw=(config.COL_METER_PRODUCTION_KW, "sum"),
            _sum_inverter_kw=(config.COL_TOTAL_INVERTER_KW, "sum"),
        )
        .reset_index()
        .rename(columns={config.COL_TIME_BUCKET: "time_bucket"})
    )

    # Energy-weighted efficiency per time bucket (not a mean of per-interval ratios).
    inv_pos = agg["_sum_inverter_kw"].where(agg["_sum_inverter_kw"] > 0)
    agg["avg_efficiency_pct"] = 100.0 * agg["_sum_meter_kw"] / inv_pos

    float_cols = ["avg_efficiency_pct", "avg_loss_delta_kw"]
    agg[float_cols] = agg[float_cols].round(3)

    bucket_order = ["Morning", "Peak", "Afternoon"]
    agg["time_bucket"] = pd.Categorical(agg["time_bucket"], categories=bucket_order, ordered=True)
    agg = agg.sort_values("time_bucket").reset_index(drop=True)

    label_map = {
        "Morning":   "Morning (6am–9am)",
        "Peak":      "Peak (10am–1pm)",
        "Afternoon": "Afternoon (2pm–5pm)",
    }
    agg["time_bucket"] = agg["time_bucket"].map(label_map)

    return pd.DataFrame({
        "Time Period":             agg["time_bucket"],
        "Valid Readings":          agg["row_count"],
        "Average Efficiency (%)":  agg["avg_efficiency_pct"],
        "Average Power Loss (kW)": agg["avg_loss_delta_kw"],
    })


def summarise_inverters(record: SiteRecord) -> dict:
    """Return a dict of inverter power share metrics.

    Equal share = 100 / n_inverters. A site is flagged if any inverter deviates
    more than INVERTER_IMBALANCE_TOLERANCE_PP percentage points from that equal share.
    """
    df = _reportable_df(record)
    valid = df[df[config.COL_TOTAL_INVERTER_KW] > 0]

    inv_cols = sorted(
        [c for c in df.columns if re.match(r"^Inverter \d+ AC kW$", c)],
        key=lambda c: int(re.search(r"\d+", c).group()),
    )

    n = len(inv_cols)
    equal_share = 100.0 / n

    row = {"Site": record.site_id}
    shares = []
    for i, col in enumerate(inv_cols, start=1):
        share = round((valid[col] / valid[config.COL_TOTAL_INVERTER_KW] * 100).mean(), 3)
        row[f"Inverter {i} Power Share (%)"] = share
        shares.append(share)

    imbalanced = any(abs(s - equal_share) > config.INVERTER_IMBALANCE_TOLERANCE_PP for s in shares)
    row["Notes"] = "Imbalance detected" if imbalanced else ""
    return row


def summarise_gap_months(record: SiteRecord) -> list:
    """Return month-level data gaps: months present in the raw data with meter and/or
    inverter activity recorded, but too few clean intervals to report an efficiency.

    A month is reported in the monthly table only if it has at least
    config.MIN_CLEAN_INTERVALS_PER_MONTH clean (efficiency-defined) intervals. Every other
    month that still had recorded data is surfaced here instead of silently vanishing, so
    "no data" is distinguishable from "data present but dropped/too sparse to trust". The
    filter logic is not touched; this only inspects what survived.

    Reasons (checked in priority order; a config exclusion wins over an anomaly, which
    wins over the count-based reasons):
      - "CT issue - meter readings unreliable" — month listed in the site's
        SITE_CONFIGS excluded_months; data exists but is not trusted for efficiency
      - "efficiency anomaly - possible meter or instrumentation fault" — the month's
        efficiency is statistically far below the site's own median (see _anomalous_periods)
      - "insufficient clean intervals" — some clean data, but below the monthly minimum
        (e.g. a handful of timing-jittered rows, or CT/comms dropouts gutting the month)
      - "incomplete inverter data" — meter and generation present, but never all inverters
        reporting at once, so zero intervals survived filter_inverter_active
      - "no inverter telemetry" / "no meter data" — only one side recorded that month

    Returns a list of dicts (chronological) with keys: month, raw_rows, meter_rows,
    gen_rows, clean_rows, reason. Months with no recorded data at all (pre-commissioning,
    true outages) are not reported — they are genuinely absent, not gaps.
    """
    raw = record.raw_df
    if raw is None or len(raw) == 0:
        return []

    period = pd.to_datetime(raw[config.COL_TIMESTAMP], errors="coerce").dt.to_period("M")
    inv_cols = [c for c in raw.columns if re.match(r"^Inverter \d+ AC kW$", c)]
    inv_total = raw[inv_cols].sum(axis=1) if inv_cols else pd.Series(0.0, index=raw.index)
    if config.COL_METER_PRODUCTION_KW in raw.columns:
        meter = pd.to_numeric(raw[config.COL_METER_PRODUCTION_KW], errors="coerce")
    else:
        meter = pd.Series(0.0, index=raw.index)

    # Count good-day, efficiency-defined intervals per month — the same basis as the monthly
    # table's row count (which is restricted to good days). A month is "reported" iff this
    # reaches the minimum threshold; a month with clean data but too few GOOD-day intervals
    # surfaces here as "insufficient clean intervals" rather than silently vanishing.
    clean = _keep_good_days(record.enriched_df, _good_day_dates(record))
    clean = clean[clean[config.COL_EFFICIENCY_PCT].notna()]
    clean_counts = (
        pd.to_datetime(clean[config.COL_TIMESTAMP], errors="coerce")
        .dt.to_period("M").value_counts()
    )
    min_clean = config.MIN_CLEAN_INTERVALS_PER_MONTH
    excluded = _excluded_periods(record.site_id)
    anomalous = _anomalous_periods(record)

    work = pd.DataFrame({"period": period, "meter": meter, "gen": inv_total})
    gaps = []
    for p, sub in work.groupby("period"):  # groupby drops NaT and sorts chronologically
        clean_count = int(clean_counts.get(p, 0))
        meter_rows = int((sub["meter"] > 0).sum())
        gen_rows = int((sub["gen"] > 0).sum())
        if p in excluded:
            reason = "CT issue - meter readings unreliable"  # manager-confirmed; overrides all
        elif p in anomalous:
            reason = "efficiency anomaly - possible meter or instrumentation fault"
        elif clean_count >= min_clean:
            continue  # enough clean data — reported as a real month in the table
        elif meter_rows == 0 and gen_rows == 0:
            continue  # genuinely no data this month — absent, not a gap
        elif clean_count > 0:
            reason = "insufficient clean intervals"  # below the monthly minimum to trust a number
        elif gen_rows == 0:
            reason = "no inverter telemetry"         # meter recording, inverters never reported
        elif meter_rows == 0:
            reason = "no meter data"                 # inverters generating, meter not recording
        else:
            reason = "incomplete inverter data"      # both present, but not all inverters at once
        gaps.append({
            "month":      f"{calendar.month_abbr[p.month]} {p.year}",
            "period":     p,
            "raw_rows":   int(len(sub)),
            "meter_rows": meter_rows,
            "gen_rows":   gen_rows,
            "clean_rows": clean_count,
            "reason":     reason,
        })
    return gaps


# Short, audience-friendly labels for the condensed Data Gaps lines, keyed by the verbose
# reason recorded in summarise_gap_months. Unmapped reasons fall back to their raw text.
_GAP_REASON_LABELS = {
    "CT issue - meter readings unreliable":                          "CT issue - meter readings unreliable",
    "efficiency anomaly - possible meter or instrumentation fault":  "efficiency anomaly",
    "insufficient clean intervals":                                  "insufficient clean intervals",
    "incomplete inverter data":                                      "incomplete inverter telemetry",
    "no inverter telemetry":                                         "no inverter telemetry",
    "no meter data":                                                 "no meter data",
}


def _condense_gap_windows(gaps: list) -> list:
    """Collapse the per-month gap list into one line per contiguous same-reason window.

    Consecutive calendar months that share a reason are merged so a long outage reads as a
    single line, e.g. 'Sep 2024 - Jul 2025 (11 months): incomplete inverter telemetry'
    instead of eleven interval-count rows. `gaps` is the chronological list from
    summarise_gap_months (each item carries a 'period' Period and a 'reason'). Returns a
    list of formatted strings in chronological order.
    """
    windows = []
    for g in gaps:
        p, reason = g["period"], g["reason"]
        if windows and windows[-1]["reason"] == reason and windows[-1]["end"] + 1 == p:
            windows[-1]["end"] = p
            windows[-1]["n"] += 1
        else:
            windows.append({"start": p, "end": p, "reason": reason, "n": 1})

    lines = []
    for w in windows:
        label = _GAP_REASON_LABELS.get(w["reason"], w["reason"])
        start = f"{calendar.month_abbr[w['start'].month]} {w['start'].year}"
        if w["n"] == 1:
            span = f"{start} (1 month)"
        else:
            end = f"{calendar.month_abbr[w['end'].month]} {w['end'].year}"
            span = f"{start} – {end} ({w['n']} months)"
        lines.append(f"{span}: {label}")
    return lines


# ── CSV output ────────────────────────────────────────────────────────────────

def write_cleaned_csv(record: SiteRecord, output_dir: Path = OUTPUT_DIR) -> Path:
    """Write enriched_df to output/<site_id>_cleaned.csv and return the path."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{record.site_id}_cleaned.csv"
    record.enriched_df.to_csv(path, index=False)
    return path


# ── Per-day diagnostic (sites that triggered the adaptive phase threshold) ──────
# Filter labels, in pipeline order, for the per-day kill attribution. Each producing interval
# is charged to the FIRST filter that removed it; the day's "primary kill filter" is whichever
# charged the most. This is purely a read-only replay of the existing filters at the site's
# chosen phase threshold — no filter logic is modified.
_DIAG_FILTER_STEPS = [
    ("Value Spikes",    "filter_value_spikes"),
    ("Inverter Active", "filter_inverter_active"),
    ("Inverter Comms",  "filter_inverter_comms"),
    ("Meter Comms",     "filter_meter_comms"),
    ("Phase Current",   "filter_phase_current"),
    ("Gross Outliers",  "filter_gross_outliers"),
]


def _daily_diagnostic_frame(record: SiteRecord) -> pd.DataFrame:
    """One row per calendar day present in the site's raw data, classifying it good/bad and
    attributing each producing interval to the filter that removed it.

    Re-runs the six filters in order at the site's chosen phase threshold (from
    cleaned_df.attrs). The filters are pure, so this never mutates the record or pipeline
    state. Producing day / good-day definitions match calculator.day_quality (the single
    source of truth): a producing interval has raw meter > NIGHTTIME_KW_THRESHOLD; a producing
    day is GOOD when at least GOOD_DAY_MIN_CLEAN_PCT of those intervals survive all six filters.
    Counts and kill attribution are over producing intervals only.
    """
    from src import cleaners

    raw = record.raw_df
    sid = record.site_id
    attrs = getattr(record.cleaned_df, "attrs", None) or {}
    threshold = attrs.get("phase_threshold_used")  # None -> filters use the global default

    ts = pd.to_datetime(raw[config.COL_TIMESTAMP], errors="coerce")
    date = ts.dt.normalize()
    meter = pd.to_numeric(raw[config.COL_METER_PRODUCTION_KW], errors="coerce")
    prod = (meter > config.NIGHTTIME_KW_THRESHOLD).to_numpy()

    # Replay the stack, charging each removed row to the first filter that dropped it.
    kill = pd.Series(pd.NA, index=raw.index, dtype="object")
    df = raw
    prev = raw.index
    sink = io.StringIO()
    with contextlib.redirect_stdout(sink):
        for label, fname in _DIAG_FILTER_STEPS:
            fn = getattr(cleaners, fname)
            df = fn(df, sid, threshold=threshold) if fname == "filter_phase_current" else fn(df, sid)
            removed = prev.difference(df.index)
            kill.loc[removed] = label
            prev = df.index

    work = pd.DataFrame({"date": date, "prod": prod, "kill": kill.to_numpy()})
    work = work[work["date"].notna()]

    rows = []
    for d, sub in work.groupby("date"):
        prod_sub = sub[sub["prod"]]
        n_prod = len(prod_sub)
        killed = prod_sub["kill"].dropna()      # producing intervals removed, labelled by filter
        n_surv = n_prod - len(killed)
        if n_prod == 0:                         # day present but no production (winter/all-night)
            rows.append({
                "Date": str(d.date()), "Good/Bad": "", "Clean Fraction": "",
                "Producing Intervals": 0, "Surviving Intervals": 0,
                "Primary Kill Filter": "", "Intervals Killed": 0,
            })
            continue
        frac = n_surv / n_prod
        if len(killed):
            vc = killed.value_counts()
            primary, n_killed = str(vc.index[0]), int(vc.iloc[0])
        else:
            primary, n_killed = "None (all survived)", 0
        rows.append({
            "Date": str(d.date()),
            "Good/Bad": "Good" if frac >= config.GOOD_DAY_MIN_CLEAN_PCT else "Bad",
            "Clean Fraction": f"{frac * 100:.1f}%",
            "Producing Intervals": n_prod,
            "Surviving Intervals": n_surv,
            "Primary Kill Filter": primary,
            "Intervals Killed": n_killed,
        })
    return pd.DataFrame(rows)


def write_daily_diagnostic(record: SiteRecord, output_dir: Path = OUTPUT_DIR) -> Path:
    """Write a formatted per-day diagnostic workbook for one site and return the path.

    Green rows = good days, red rows = bad days, header frozen, columns auto-fit. Intended for
    sites that triggered the adaptive phase threshold (see run_all_reports).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    frame = _daily_diagnostic_frame(record)
    headers = ["Date", "Good/Bad", "Clean Fraction", "Producing Intervals",
               "Surviving Intervals", "Primary Kill Filter", "Intervals Killed"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Diagnostic"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    green = PatternFill("solid", fgColor="C6EFCE")  # Excel "good" green
    red = PatternFill("solid", fgColor="FFC7CE")    # Excel "bad" red
    for _, r in frame.iterrows():
        ws.append([r[h] for h in headers])
        fill = green if r["Good/Bad"] == "Good" else red if r["Good/Bad"] == "Bad" else None
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    for col_cells in ws.columns:
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = longest + 2

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{record.site_id}_daily_diagnostic.xlsx"
    wb.save(path)
    return path


# ── Orchestrator ──────────────────────────────────────────────────────────────

def _phase_sensitivity(record: SiteRecord) -> list:
    """Return avg efficiency for phase thresholds 1–5%, running the full filter+calc pipeline.

    The phase imbalance filter is reproduced inline here so cleaners.py stays unchanged.
    """
    from src.cleaners import filter_inverter_active, filter_gross_outliers
    from src.calculator import run_all_calculations

    def _apply_phase_filter(df: pd.DataFrame, threshold: float) -> pd.DataFrame:
        lower_map = {c.lower(): c for c in df.columns}
        flag = pd.Series(False, index=df.index)
        for patterns in (config.ACE_METER_VOLTAGE_PATTERNS, config.ACE_METER_CURRENT_PATTERNS):
            present = []
            for pat in patterns:
                match = next((orig for lc, orig in lower_map.items() if pat.lower() in lc), None)
                if match:
                    present.append(match)
            if len(present) < 2:
                continue
            vals = df[present]
            mean = vals.mean(axis=1)
            devs = vals.sub(mean, axis=0).abs().div(mean.where(mean != 0), axis=0)
            flag |= (devs > threshold).any(axis=1).fillna(False)
        return df[~flag].copy()

    nonreportable = _nonreportable_periods(record)
    sink = io.StringIO()
    rows = []
    for pct in range(1, 6):
        threshold = pct / 100
        with contextlib.redirect_stdout(sink):
            df = _apply_phase_filter(record.raw_df, threshold)
            df = filter_inverter_active(df, record.site_id)
            df = filter_gross_outliers(df, record.site_id)
            enriched = run_all_calculations(df)
        if nonreportable:  # keep the sensitivity figures consistent with the headline numbers
            per = pd.to_datetime(enriched[config.COL_TIMESTAMP], errors="coerce").dt.to_period("M")
            enriched = enriched[~per.isin(nonreportable)]
        rows.append({
            "threshold_pct":     pct,
            "rows_kept":         len(df),
            "avg_efficiency_pct": round(_weighted_efficiency_pct(enriched), 2),
        })
    return rows


def _print_sensitivity_table(all_sens: dict) -> None:
    """Print a threshold × site sensitivity table using tabulate.

    all_sens: {site_name: [{"threshold_pct": int, "avg_efficiency_pct": float}, ...]}
    """
    sites = list(all_sens.keys())
    rows = []
    for pct in range(1, 6):
        row = [f"{pct}%"]
        for site in sites:
            entry = next((r for r in all_sens[site] if r["threshold_pct"] == pct), None)
            eff = entry["avg_efficiency_pct"] if entry else float("nan")
            row.append(f"{eff:.2f}%" if pd.notna(eff) else "nan%")
        rows.append(row)
    headers = ["Threshold"] + sites
    print(f"  Phase Imbalance Sensitivity (All Sites)")
    print(tabulate(rows, headers=headers, tablefmt="simple", colalign=("right",) + ("right",) * len(sites)))
    print()


def _print_summary_table(summaries: list) -> None:
    """Print a one-row-per-site summary table. The Phase Threshold column shows the value the
    adaptive search actually used for each site, with a '*' when it was auto-adjusted from the
    global default (more than one iteration); a footnote is shown when any '*' appears."""
    width = 103
    print(f"\n{'='*width}")
    print(f"  Summary")
    print(f"{'='*width}")
    print(f"  {'Site':<22}  {'Date Range':<23}  {'Phase Threshold':>15}  "
          f"{'Avg Efficiency':>16}  {'Total Energy Lost':>20}")
    print(f"  {'-'*22}  {'-'*23}  {'-'*15}  {'-'*16}  {'-'*20}")
    any_auto = False
    for s in summaries:
        date_range   = (
            f"{s['date_range_start']} – {s['date_range_end']}"
            if s["date_range_start"] else "N/A"
        )
        used = s.get("phase_threshold_used")
        if used is None:
            phase_str = f"{'N/A':>15}"
        else:
            mark = " *" if s.get("phase_threshold_auto") else ""
            any_auto = any_auto or bool(mark)
            phase_str = f"{_fmt_thr(used) + mark:>15}"
        avg_eff      = s["avg_efficiency_pct"]
        total_energy = s["total_energy_lost_kwh"]
        eff_str    = f"{avg_eff:>15.2f}%"         if pd.notna(avg_eff)      else f"{'N/A':>15} "
        energy_str = f"{total_energy:>16.1f} kWh" if pd.notna(total_energy) else f"{'N/A':>16}    "
        print(f"  {s['site_name']:<22}  {date_range:<23}  {phase_str}  {eff_str}  {energy_str}")
    print(f"{'='*width}")
    if any_auto:
        print("  * phase threshold auto-adjusted from default")
    print()


def _fmt_thr(x: float) -> str:
    """Format a phase threshold for display. Two decimals for the normal config (step 0.01),
    extending precision only if two decimals would misrepresent a finer custom step."""
    s = f"{x:.2f}"
    if abs(float(s) - x) > 1e-9:
        s = f"{x:.4f}".rstrip("0").rstrip(".")
    return s


def _phase_threshold_note(record: SiteRecord) -> str:
    """One-line description of the per-site adaptive phase threshold, built from the metadata
    cleaners.run_all_filters stored on cleaned_df.attrs. Returns '' if no metadata is present.

    - Reached the good-day floor without adjusting:  'Phase threshold 0.04 (1 iteration)'
    - Stepped up and reached it:  'Phase threshold auto-adjusted to 0.10 (7 iterations)'
    - Hit the ceiling first:  'Phase threshold reached ceiling 0.30 — only 5 good days found'
    """
    attrs = getattr(record.cleaned_df, "attrs", None) or {}
    if "phase_threshold_used" not in attrs:
        return ""
    used    = _fmt_thr(attrs["phase_threshold_used"])
    iters   = int(attrs.get("phase_threshold_iterations", 1))
    it_word = "iteration" if iters == 1 else "iterations"
    if attrs.get("phase_threshold_hit_ceiling"):
        ceiling = _fmt_thr(attrs.get("phase_threshold_ceiling", attrs["phase_threshold_used"]))
        good    = int(attrs.get("phase_threshold_good_days", 0))
        return f"Phase threshold reached ceiling {ceiling} — only {good} good days found"
    if iters > 1:
        return f"Phase threshold auto-adjusted to {used} ({iters} {it_word})"
    return f"Phase threshold {used} ({iters} {it_word})"


def run_all_reports(records: List[SiteRecord], output_dir: Path = OUTPUT_DIR) -> None:
    """Print the cross-site Summary first, then per-site tables, and write cleaned CSVs."""
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # allow the ≥ glyph on Windows consoles
    except Exception:
        pass

    if not records:
        print("[reporter] No records loaded — nothing to report.")
        return

    # Compute every site's summary and write its cleaned CSV up front (silently), so the
    # cross-site Summary table can print FIRST, before any per-site detail.
    all_summaries = []
    for r in records:
        all_summaries.append(summarise_site(r))
        write_cleaned_csv(r, output_dir)
        # Sites that triggered the adaptive phase threshold (more than one iteration) get a
        # per-day diagnostic workbook. Isolated so an export error can't break the report.
        if r.summary.get("phase_threshold_auto"):
            try:
                write_daily_diagnostic(r, output_dir)
            except Exception as exc:  # pragma: no cover - diagnostic is best-effort
                print(f"[reporter] could not write daily diagnostic for {r.site_id}: {exc}")

    _print_summary_table(all_summaries)

    all_sens: dict = {}
    for r in records:
        summary = r.summary
        monthly = summarise_by_month(r)

        # ── Site header ───────────────────────────────────────────────────────
        bar = "=" * 70
        print(f"\n{bar}")
        print(f"  {summary['site_name']}")
        print(bar)

        # ── Adaptive phase-threshold note (what the search landed on for this site) ──
        note = _phase_threshold_note(r)
        if note:
            print(f"  {note}")
            print()

        # ── Monthly efficiency (good days only) ───────────────────────────────
        print(f"  {'Month':<10}  {'Good/Bad Days':>13}  {'Avg Efficiency':>16}  {'Total Energy Lost':>20}")
        print(f"  {'-'*10}  {'-'*13}  {'-'*16}  {'-'*20}")
        for _, row in monthly.iterrows():
            eff    = row["Average Efficiency (%)"]
            energy = row["Total Energy Lost (kWh)"]
            gb     = f"{int(row['Good Days'])}/{int(row['Bad Days'])}"
            eff_str    = f"{eff:>15.2f}%"       if pd.notna(eff)    else f"{'N/A':>15} "
            energy_str = f"{energy:>16.1f} kWh" if pd.notna(energy) else f"{'N/A':>16}    "
            print(f"  {row['Month']:<10}  {gb:>13}  {eff_str}  {energy_str}")
        print(bar)
        avg_eff      = summary["avg_efficiency_pct"]
        total_energy = summary["total_energy_lost_kwh"]
        gb_total     = f"{summary['good_days']}/{summary['bad_days']}"
        eff_str      = f"{avg_eff:>15.2f}%"         if pd.notna(avg_eff)      else f"{'N/A':>15} "
        energy_str   = f"{total_energy:>16.1f} kWh" if pd.notna(total_energy) else f"{'N/A':>16}    "
        print(f"  {'OVERALL':<10}  {gb_total:>13}  {eff_str}  {energy_str}")
        print(f"{bar}\n")

        if VERBOSE:
            all_sens[r.site_id] = _phase_sensitivity(r)

    if VERBOSE:
        _print_sensitivity_table(all_sens)
